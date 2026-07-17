#!/usr/bin/env python3
"""
update_site_data.py — keeps the live site in sync with two source-of-truth files:

  data/scholar.pdf   -> Measures counters: citations, h-index, i10-index,
                        plus the "as of" month/year next to the Scholar link.
                        (Print/save the Google Scholar profile page as PDF and drop it here.)
  data/patents.xlsx  -> The Patents section rows (sheet "patents"), the Patents
                        counter (= row count), and the publications counter
                        (sheet "config", key "publications" — Scholar's PDF
                        doesn't print a total, so it lives here).

Idempotent: running twice in a row produces no further changes.
Run from anywhere:  python scripts/update_site_data.py
Test against copies: python scripts/update_site_data.py --scholar X.pdf --patents Y.xlsx --index Z.html
"""

import argparse
import html as htmllib
import re
import sys
from pathlib import Path

from pypdf import PdfReader
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent

MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]


# ----------------------------------------------------------------------------
# parsers
# ----------------------------------------------------------------------------
def parse_scholar(pdf_path: Path) -> dict:
    text = "\n".join((page.extract_text() or "") for page in PdfReader(str(pdf_path)).pages)

    def grab(pattern, label):
        m = re.search(pattern, text)
        if not m:
            sys.exit(f"ERROR: could not find {label} in {pdf_path} — is this a Scholar profile PDF?")
        return int(m.group(1).replace(",", ""))

    out = {
        "citations": grab(r"Citations\s+([\d,]+)", "Citations"),
        "hindex":    grab(r"h-index\s+(\d+)", "h-index"),
        "i10":       grab(r"i10-index\s+(\d+)", "i10-index"),
    }

    # "as of" from the PDF's own print timestamp, e.g. "7/16/26, 7:05 PM"
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2}),", text)
    if m:
        month, year2 = int(m.group(1)), int(m.group(3))
        out["asof"] = f"{MONTHS[month - 1]} 20{year2:02d}"
    return out


def parse_patents(xlsx_path: Path) -> tuple[list[dict], dict]:
    wb = load_workbook(str(xlsx_path), data_only=True)
    if "patents" not in wb.sheetnames:
        sys.exit(f"ERROR: {xlsx_path} has no 'patents' sheet.")
    ws = wb["patents"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, raw))
        if not row.get("title"):
            continue
        rows.append(row)
    rows.sort(key=lambda r: (r.get("order") is None, r.get("order")))

    config = {}
    if "config" in wb.sheetnames:
        for k, v in wb["config"].iter_rows(min_row=2, values_only=True):
            if k and not str(k).startswith("#") and v is not None:
                config[str(k).strip()] = v
    return rows, config


# ----------------------------------------------------------------------------
# renderers / injectors
# ----------------------------------------------------------------------------
def esc(v) -> str:
    return htmllib.escape(str(v).strip()) if v is not None else ""


def render_patent_rows(rows: list[dict]) -> str:
    out = []
    for i, r in enumerate(rows, start=1):
        status = esc(r.get("status", ""))
        granted = status.lower().startswith("grant")
        status_badge = (f'<em class="aw">&#10022; Granted</em>' if granted
                        else f"<em>{status}</em>")
        badges = [status_badge]
        if r.get("assignee"):
            badges.append(f"<em>{esc(r['assignee'])}</em>")
        for b in str(r.get("extra_badges") or "").split("|"):
            if b.strip():
                badges.append(f"<em>{esc(b)}</em>")

        verb = status.split()[0] if status else "Filed"
        meta_lines = f"{verb} {esc(r.get('status_date',''))}<br>Filed {esc(r.get('filed_date',''))}"

        out.append(f'''    <a class="patrow" href="{esc(r.get("link","#"))}" target="_blank" rel="noopener" data-reveal>
      <span class="r-idx">{i:02d}</span>
      <span class="r-main">
        <h3>{esc(r["title"])}</h3>
        <span class="r-sub">{esc(r.get("description",""))}</span>
        <span class="r-badges">{"".join(badges)}</span>
      </span>
      <span class="r-meta"><span class="r-cites">{esc(r.get("number",""))}</span>{meta_lines}</span>
    </a>''')
    return "\n".join(out)


def set_metric(html: str, metric: str, value: int) -> str:
    pattern = rf'(data-metric="{metric}" data-count=")\d+(")'
    new, n = re.subn(pattern, rf"\g<1>{value}\g<2>", html)
    if n != 1:
        sys.exit(f"ERROR: expected exactly 1 counter for metric '{metric}', found {n}.")
    return new


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--scholar", default=str(ROOT / "data" / "scholar.pdf"))
    ap.add_argument("--patents", default=str(ROOT / "data" / "patents.xlsx"))
    ap.add_argument("--index",   default=str(ROOT / "index.html"))
    args = ap.parse_args()

    scholar = parse_scholar(Path(args.scholar))
    patents, config = parse_patents(Path(args.patents))

    index_path = Path(args.index)
    html = index_path.read_text(encoding="utf-8")
    original = html

    html = set_metric(html, "citations", scholar["citations"])
    html = set_metric(html, "hindex", scholar["hindex"])
    html = set_metric(html, "i10", scholar["i10"])
    html = set_metric(html, "patents", len(patents))
    if "publications" in config:
        html = set_metric(html, "publications", int(config["publications"]))

    if scholar.get("asof"):
        html, n = re.subn(r'(<span data-scholar-asof>)[^<]*(</span>)',
                          rf"\g<1>{scholar['asof']}\g<2>", html)
        if n != 1:
            sys.exit("ERROR: data-scholar-asof marker not found exactly once.")

    block = "<!-- PATENTS:START -->\n" + render_patent_rows(patents) + "\n<!-- PATENTS:END -->"
    html, n = re.subn(r"<!-- PATENTS:START -->.*?<!-- PATENTS:END -->", block, html, flags=re.S)
    if n != 1:
        sys.exit("ERROR: PATENTS markers not found exactly once.")

    if html != original:
        index_path.write_text(html, encoding="utf-8")
        print(f"UPDATED {index_path.name}: citations={scholar['citations']} "
              f"h={scholar['hindex']} i10={scholar['i10']} patents={len(patents)} "
              f"publications={config.get('publications','(unchanged)')} asof={scholar.get('asof','(unchanged)')}")
    else:
        print("No changes — site already in sync with data/ sources.")


if __name__ == "__main__":
    main()
